from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


for row in ws.iter_rows(min_row=2):
    #번호0, 영어1, 수학2
    if int(row[1].value) > 80: # 1번째 값이 80이상인 값이면~~
        print(row[0].value,"번 학생은 영어 천재") # 0번째 값을 프린트 하면 몇번 학생~~~

for row in ws.iter_rows(max_row=1): # cell 값 영어를  컴퓨터로 바꾸는 방법
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")

