# 수식,  01:46:27
import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()# 오늘 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)" # 1,2,3을 더한 6( 합계)
ws["A3"] = "=AVERAGE(1, 2, 3)" #평균

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"# 30
ws["A7"] = "=a4*a5" # 200


wb.save("sample_formula.xlsx")