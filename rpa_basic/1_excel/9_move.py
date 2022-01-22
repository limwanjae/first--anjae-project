from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# (번호, 영어, 수학) 데이터를
# (번호, 국어, 영어, 수학) 하기 위해서 (영어, 수학)을 이동시키는 방법

ws.move_range("B1:C11", rows=0, cols=1)# row는 그대로, 열은 1칸
ws["B1"].value = "국어" # B1 셀에 "국어" 입력

wb.save("sample_korean.xlsx")
