from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()# 새로운 sheet를 기본 이름으로 생성
ws.title ="MySheet"# sheet 이름 변경
ws.sheet_properties.tabColor ="ff66ff"# RGB형태로

# 0sheet, 1MySheet, 2YourSheet
ws1 = wb.create_sheet("YourSheet")# 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet",2)# 2번째 index에 시트를 생성

new_ws = wb["NewSheet"] # Dict형태로 Sheet에 접근

print(wb.sheetnames)# 모든 Sheet 이름 확인

# Sheet복사
new_ws["A1"] = "Test" # A1셀에 Test입력
target = wb.copy_worksheet(new_ws) # new_ws시트를 카피해서 target 변수에 저장
target.title = "Copied Sheet" # target시트의 제목을  Copied Sheet로 바꾼다

wb.save("sample.xlsx")