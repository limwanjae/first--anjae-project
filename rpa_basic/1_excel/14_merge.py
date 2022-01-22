# 셀 병합  01:55:34
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("b2:d2") #b2 부터 d2 까지 합침
ws["b2"].value = "merged cell"



wb.save("sample_merge.xlsx")

